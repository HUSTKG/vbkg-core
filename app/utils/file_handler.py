# app/utils/file_handler.py

import asyncio
import logging
import mimetypes
import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

# PDF processing
try:
    import pdfplumber
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Office documents
try:
    import docx
    from openpyxl import load_workbook
    OFFICE_AVAILABLE = True
except ImportError:
    OFFICE_AVAILABLE = False

import csv
# Text processing
import json
import xml.etree.ElementTree as ET

from app.core.config import settings

logger = logging.getLogger(__name__)


class FileProcessor:
    """Comprehensive file processing for various formats"""
    
    def __init__(self):
        self.supported_formats = {
            # Text formats
            'text/plain': self._process_text,
            'text/csv': self._process_csv,
            'application/json': self._process_json,
            'application/xml': self._process_xml,
            'text/xml': self._process_xml,
            
            # PDF formats
            'application/pdf': self._process_pdf,
            
            # Office formats
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': self._process_docx,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': self._process_xlsx,
            'application/msword': self._process_doc,
            
            # Web formats
            'text/html': self._process_html,
            'text/markdown': self._process_markdown,
        }
    
    async def process_file(
        self, 
        file_path: str, 
        file_type: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Process file and extract text content with metadata
        
        Args:
            file_path: Path to the file
            file_type: MIME type of the file (auto-detected if None)
            
        Returns:
            Dictionary containing extracted content and metadata
        """
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Auto-detect file type if not provided
        if not file_type:
            file_type = self._detect_file_type(file_path)
        
        # Check if format is supported
        if file_type not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_type}")
        
        # Get file info
        file_info = self._get_file_info(file_path)
        
        try:
            # Process file based on type
            processor = self.supported_formats[file_type]
            content_data = await processor(file_path)
            
            return {
                "success": True,
                "file_info": file_info,
                "content": content_data,
                "file_type": file_type,
                "processing_method": processor.__name__
            }
            
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}")
            return {
                "success": False,
                "file_info": file_info,
                "error": str(e),
                "file_type": file_type
            }
    
    def _detect_file_type(self, file_path: str) -> str:
        """Detect file MIME type"""
        
        mime_type, _ = mimetypes.guess_type(file_path)
        
        if mime_type:
            return mime_type
        
        # Fallback to extension-based detection
        extension = Path(file_path).suffix.lower()
        extension_map = {
            '.txt': 'text/plain',
            '.csv': 'text/csv',
            '.json': 'application/json',
            '.xml': 'application/xml',
            '.pdf': 'application/pdf',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.doc': 'application/msword',
            '.html': 'text/html',
            '.htm': 'text/html',
            '.md': 'text/markdown',
        }
        
        return extension_map.get(extension, 'application/octet-stream')
    
    def _get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get file metadata"""
        
        stat = os.stat(file_path)
        
        return {
            "name": os.path.basename(file_path),
            "size": stat.st_size,
            "created": stat.st_ctime,
            "modified": stat.st_mtime,
            "extension": Path(file_path).suffix.lower()
        }
    
    # =============================================
    # TEXT FORMAT PROCESSORS
    # =============================================
    
    async def _process_text(self, file_path: str) -> Dict[str, Any]:
        """Process plain text file"""
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        except UnicodeDecodeError:
            # Try with different encoding
            with open(file_path, 'r', encoding='latin-1') as file:
                content = file.read()
        
        return {
            "text": content,
            "char_count": len(content),
            "line_count": content.count('\n') + 1,
            "word_count": len(content.split())
        }
    
    async def _process_csv(self, file_path: str) -> Dict[str, Any]:
        """Process CSV file"""
        
        rows = []
        headers = []
        
        with open(file_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            
            # Get headers
            headers = next(reader, [])
            
            # Read rows (limit for safety)
            for i, row in enumerate(reader):
                if i >= 10000:  # Safety limit
                    break
                rows.append(row)
        
        # Convert to text for entity extraction
        text_content = []
        
        # Add headers as context
        if headers:
            text_content.append("Columns: " + ", ".join(headers))
        
        # Add sample rows as text
        for row in rows[:100]:  # First 100 rows
            if len(row) >= len(headers):
                row_text = " | ".join(f"{headers[i] if i < len(headers) else f'col_{i}'}: {cell}" 
                                     for i, cell in enumerate(row) if cell.strip())
                if row_text:
                    text_content.append(row_text)
        
        return {
            "text": "\n".join(text_content),
            "headers": headers,
            "row_count": len(rows),
            "column_count": len(headers),
            "sample_rows": rows[:10]
        }
    
    async def _process_json(self, file_path: str) -> Dict[str, Any]:
        """Process JSON file"""
        
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        # Convert JSON structure to readable text
        text_content = self._json_to_text(data)
        
        return {
            "text": text_content,
            "json_data": data,
            "structure_type": type(data).__name__
        }
    
    def _json_to_text(self, obj: Any, prefix: str = "") -> str:
        """Convert JSON object to readable text"""
        
        text_parts = []
        
        if isinstance(obj, dict):
            for key, value in obj.items():
                if isinstance(value, (dict, list)):
                    text_parts.append(f"{prefix}{key}:")
                    text_parts.append(self._json_to_text(value, prefix + "  "))
                else:
                    text_parts.append(f"{prefix}{key}: {value}")
        
        elif isinstance(obj, list):
            for i, item in enumerate(obj[:50]):  # Limit for safety
                if isinstance(item, (dict, list)):
                    text_parts.append(f"{prefix}Item {i}:")
                    text_parts.append(self._json_to_text(item, prefix + "  "))
                else:
                    text_parts.append(f"{prefix}Item {i}: {item}")
        
        else:
            text_parts.append(f"{prefix}{obj}")
        
        return "\n".join(text_parts)
    
    async def _process_xml(self, file_path: str) -> Dict[str, Any]:
        """Process XML file"""
        
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Extract text content
            text_content = self._xml_to_text(root)
            
            return {
                "text": text_content,
                "root_tag": root.tag,
                "element_count": len(list(root.iter()))
            }
            
        except ET.ParseError as e:
            # Try to read as text if XML parsing fails
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            return {
                "text": content,
                "error": f"XML parsing failed: {e}",
                "fallback": "text"
            }
    
    def _xml_to_text(self, element: ET.Element, prefix: str = "") -> str:
        """Convert XML element to readable text"""
        
        text_parts = []
        
        # Element name and attributes
        if element.attrib:
            attrs = " ".join(f"{k}={v}" for k, v in element.attrib.items())
            text_parts.append(f"{prefix}{element.tag} ({attrs})")
        else:
            text_parts.append(f"{prefix}{element.tag}")
        
        # Element text
        if element.text and element.text.strip():
            text_parts.append(f"{prefix}  {element.text.strip()}")
        
        # Child elements
        for child in element:
            text_parts.append(self._xml_to_text(child, prefix + "  "))
        
        return "\n".join(text_parts)
    
    # =============================================
    # PDF PROCESSORS
    # =============================================
    
    async def _process_pdf(self, file_path: str) -> Dict[str, Any]:
        """Process PDF file"""
        
        if not PDF_AVAILABLE:
            raise ImportError("PDF processing libraries not available. Install PyPDF2 and pdfplumber.")
        
        text_content = ""
        page_count = 0
        metadata = {}
        
        try:
            # Try pdfplumber first (better text extraction)
            with pdfplumber.open(file_path) as pdf:
                page_count = len(pdf.pages)
                
                text_parts = []
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
                
                text_content = "\n\n".join(text_parts)
                
                # Extract metadata
                if pdf.metadata:
                    metadata = {
                        "title": pdf.metadata.get("Title"),
                        "author": pdf.metadata.get("Author"),
                        "subject": pdf.metadata.get("Subject"),
                        "creator": pdf.metadata.get("Creator")
                    }
        
        except Exception as e:
            # Fallback to PyPDF2
            logger.warning(f"pdfplumber failed, trying PyPDF2: {e}")
            
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                page_count = len(pdf_reader.pages)
                
                text_parts = []
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
                
                text_content = "\n\n".join(text_parts)
                
                # Extract metadata
                if pdf_reader.metadata:
                    metadata = {
                        "title": pdf_reader.metadata.get("/Title"),
                        "author": pdf_reader.metadata.get("/Author"),
                        "subject": pdf_reader.metadata.get("/Subject"),
                        "creator": pdf_reader.metadata.get("/Creator")
                    }
        
        return {
            "text": text_content,
            "page_count": page_count,
            "char_count": len(text_content),
            "word_count": len(text_content.split()),
            "metadata": metadata
        }
    
    # =============================================
    # OFFICE DOCUMENT PROCESSORS
    # =============================================
    
    async def _process_docx(self, file_path: str) -> Dict[str, Any]:
        """Process Word document"""
        
        if not OFFICE_AVAILABLE:
            raise ImportError("Office processing libraries not available. Install python-docx.")
        
        doc = docx.Document(file_path)
        
        # Extract paragraphs
        paragraphs = []
        for para in doc.paragraphs:
            if para.text.strip():
                paragraphs.append(para.text.strip())
        
        text_content = "\n\n".join(paragraphs)
        
        # Extract metadata
        metadata = {}
        if doc.core_properties:
            metadata = {
                "title": doc.core_properties.title,
                "author": doc.core_properties.author,
                "subject": doc.core_properties.subject,
                "created": str(doc.core_properties.created) if doc.core_properties.created else None,
                "modified": str(doc.core_properties.modified) if doc.core_properties.modified else None
            }
        
        return {
            "text": text_content,
            "paragraph_count": len(paragraphs),
            "char_count": len(text_content),
            "word_count": len(text_content.split()),
            "metadata": metadata
        }
    
    async def _process_xlsx(self, file_path: str) -> Dict[str, Any]:
        """Process Excel spreadsheet"""
        
        if not OFFICE_AVAILABLE:
            raise ImportError("Office processing libraries not available. Install openpyxl.")
        
        workbook = load_workbook(file_path, read_only=True)
        
        text_parts = []
        sheet_info = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_info[sheet_name] = {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            }
            
            # Add sheet name as context
            text_parts.append(f"Sheet: {sheet_name}")
            
            # Read data (limit for safety)
            rows_read = 0
            for row in sheet.iter_rows(max_row=min(100, sheet.max_row), values_only=True):
                if rows_read >= 100:
                    break
                
                # Filter out empty cells and create text
                cell_values = [str(cell) for cell in row if cell is not None]
                if cell_values:
                    text_parts.append(" | ".join(cell_values))
                    rows_read += 1
            
            text_parts.append("")  # Separator between sheets
        
        text_content = "\n".join(text_parts)
        
        return {
            "text": text_content,
            "sheet_count": len(workbook.sheetnames),
            "sheet_info": sheet_info,
            "char_count": len(text_content),
            "word_count": len(text_content.split())
        }
    
    async def _process_doc(self, file_path: str) -> Dict[str, Any]:
        """Process legacy Word document"""
        
        # For .doc files, we'd need python-docx2txt or similar
        # For now, return an error message
        return {
            "text": "",
            "error": "Legacy .doc format not supported. Please convert to .docx",
            "suggestion": "Convert to .docx format for processing"
        }
    
    # =============================================
    # WEB FORMAT PROCESSORS
    # =============================================
    
    async def _process_html(self, file_path: str) -> Dict[str, Any]:
        """Process HTML file"""
        
        try:
            from bs4 import BeautifulSoup
            
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Extract text content
            text_content = soup.get_text(separator='\n', strip=True)
            
            # Extract metadata
            metadata = {}
            title_tag = soup.find('title')
            if title_tag:
                metadata['title'] = title_tag.get_text(strip=True)
            
            meta_tags = soup.find_all('meta')
            for meta in meta_tags:
                if meta.get('name') and meta.get('content'):
                    metadata[meta.get('name')] = meta.get('content')
            
            return {
                "text": text_content,
                "html_length": len(html_content),
                "text_length": len(text_content),
                "metadata": metadata
            }
            
        except ImportError:
            # Fallback without BeautifulSoup
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            
            # Simple HTML tag removal
            import re
            text_content = re.sub(r'<[^>]+>', ' ', html_content)
            text_content = ' '.join(text_content.split())
            
            return {
                "text": text_content,
                "html_length": len(html_content),
                "text_length": len(text_content),
                "processing_method": "simple_regex"
            }
    
    async def _process_markdown(self, file_path: str) -> Dict[str, Any]:
        """Process Markdown file"""
        
        with open(file_path, 'r', encoding='utf-8') as file:
            markdown_content = file.read()
        
        try:
            import markdown

            # Convert to HTML then extract text
            html = markdown.markdown(markdown_content)
            
            # Simple HTML tag removal for text extraction
            import re
            text_content = re.sub(r'<[^>]+>', ' ', html)
            text_content = ' '.join(text_content.split())
            
            return {
                "text": text_content,
                "markdown_length": len(markdown_content),
                "text_length": len(text_content),
                "html": html
            }
            
        except ImportError:
            # Simple markdown processing without library
            import re

            # Remove markdown syntax for basic text extraction
            text_content = re.sub(r'[#*`\[\]()_~]', '', markdown_content)
            text_content = ' '.join(text_content.split())
            
            return {
                "text": text_content,
                "markdown_length": len(markdown_content),
                "text_length": len(text_content),
                "processing_method": "simple_regex"
            }


# =============================================
# CONVENIENCE FUNCTIONS
# =============================================

# Global processor instance
_file_processor: Optional[FileProcessor] = None


def get_file_processor() -> FileProcessor:
    """Get global file processor instance"""
    
    global _file_processor
    
    if _file_processor is None:
        _file_processor = FileProcessor()
    
    return _file_processor


async def process_file(file_path: str, file_type: Optional[str] = None) -> Dict[str, Any]:
    """Process file using global processor"""
    
    processor = get_file_processor()
    return await processor.process_file(file_path, file_type)


def get_file_type(file_path: str, provided_type: Optional[str] = None) -> str:
    """Get file type with fallback detection"""
    
    if provided_type:
        return provided_type
    
    processor = get_file_processor()
    return processor._detect_file_type(file_path)


async def parse_file_content(file_path: str, file_type: str) -> str:
    """Extract text content from file"""
    
    result = await process_file(file_path, file_type)
    
    if result.get("success"):
        return result["content"].get("text", "")
    else:
        raise Exception(f"Failed to parse file: {result.get('error', 'Unknown error')}")


# =============================================
# FILE UTILITIES
# =============================================

async def save_upload_file(upload_file, destination_path: str) -> str:
    """Save uploaded file to destination"""
    
    os.makedirs(os.path.dirname(destination_path), exist_ok=True)
    
    with open(destination_path, "wb") as buffer:
        content = await upload_file.read()
        buffer.write(content)
    
    return destination_path


async def create_temp_file(content: bytes, suffix: str = "") -> str:
    """Create temporary file with content"""
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
        temp_file.write(content)
        return temp_file.name


def cleanup_temp_file(file_path: str):
    """Remove temporary file safely"""
    
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        logger.warning(f"Failed to cleanup temp file {file_path}: {e}")


# =============================================
# TEXT PROCESSING UTILITIES
# =============================================

def chunk_text(
    text: str, 
    chunk_size: int = 1000, 
    overlap: int = 200,
    preserve_sentences: bool = True
) -> List[str]:
    """Split text into overlapping chunks"""
    
    if not text:
        return []
    
    if len(text) <= chunk_size:
        return [text]
    
    chunks = []
    start = 0
    
    while start < len(text):
        end = start + chunk_size
        
        if preserve_sentences and end < len(text):
            # Try to break at sentence boundary
            sentence_ends = ['.', '!', '?', '\n']
            best_break = end
            
            # Look backwards for sentence boundary
            for i in range(end - 100, min(end + 100, len(text))):
                if i > 0 and text[i-1] in sentence_ends and text[i].isspace():
                    best_break = i
                    break
            
            end = best_break
        
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        
        start = end - overlap
        
        if start >= len(text):
            break
    
    return chunks


def extract_metadata_from_text(text: str) -> Dict[str, Any]:
    """Extract basic metadata from text content"""
    
    if not text:
        return {}
    
    lines = text.split('\n')
    words = text.split()
    
    return {
        "char_count": len(text),
        "word_count": len(words),
        "line_count": len(lines),
        "paragraph_count": len([line for line in lines if line.strip()]),
        "avg_words_per_line": len(words) / max(len(lines), 1),
        "reading_time_minutes": len(words) / 200  # Assuming 200 WPM reading speed
    }


if __name__ == "__main__":
    # Test file processing
    async def test_processing():
        processor = get_file_processor()
        
        # Test with a sample text file
        test_content = "This is a test document.\nIt contains multiple lines.\nFor testing purposes."
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as temp_file:
            temp_file.write(test_content)
            temp_file_path = temp_file.name
        
        try:
            result = await processor.process_file(temp_file_path)
            print(f"Processing result: {result}")
            
            # Test chunking
            chunks = chunk_text(test_content, chunk_size=20, overlap=5)
            print(f"Text chunks: {chunks}")
            
        finally:
            cleanup_temp_file(temp_file_path)
    
    asyncio.run(test_processing())
